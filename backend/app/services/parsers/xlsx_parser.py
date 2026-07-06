import io
import openpyxl
from app.services.parsers.base import DocumentParser, ExtractionResult, ExtractedPage
from app.utils.logging import get_logger

logger = get_logger(__name__)

class XlsxParser(DocumentParser):
    def parse(self, content: bytes, filename: str) -> ExtractionResult:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open XLSX '{filename}': {exc}") from exc

        pages: list[ExtractedPage] = []
        cursor = 0
        total_pages = len(wb.sheetnames)
        
        for i, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            sheet_text = [f"Sheet: {sheet_name}"]
            
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to string
                row_values = [str(cell) for cell in row if cell is not None]
                if row_values:
                    sheet_text.append(" | ".join(row_values))
                    
            page_text = "\n".join(sheet_text).strip()
            # If a sheet only contains its name but no data, we can still include it, but let's check length
            if len(sheet_text) <= 1:
                continue

            start = cursor
            end = start + len(page_text)
            
            pages.append(ExtractedPage(
                page_number=i + 1,
                text=page_text,
                char_start=start,
                char_end=end
            ))
            cursor = end + 2

        if not pages:
            raise ValueError(f"XLSX '{filename}' contains no extractable text.")
            
        full_text = "\n\n".join(p.text for p in pages)
        
        return ExtractionResult(
            pages=pages,
            full_text=full_text,
            page_count=total_pages,
            char_count=len(full_text),
            file_type="xlsx",
            parser_used="openpyxl",
            ocr_used=False,
            ocr_engine=None,
            extraction_method="native"
        )
